"""
storage.py — Excel database management using openpyxl.

Creates daily Excel files, writes formatted headers, and appends
data rows efficiently without rewriting the entire file.

The Excel schema has two sections:
  1. Raw data columns (Date-Time, Price, 9 × timeframe signals)
  2. Strategy columns  (Signal, Bias, Confidence, Entry, SL, TP, Risk, Lots, Reason)
"""

from __future__ import annotations

import os
import shutil
import tempfile
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import (
    EXCEL_COLUMNS,
    STRATEGY_COLUMNS,
    ALL_EXCEL_COLUMNS,
    DATA_DIR,
    get_excel_filename,
)
from utils.logger import get_logger

log = get_logger("storage")

# ── Header styling ────────────────────────────────────────────────────
_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
_HEADER_BORDER = Border(
    bottom=Side(style="thin", color="000000"),
    right=Side(style="thin", color="D9D9D9"),
)

# Strategy header uses a distinct colour to visually separate sections
_STRATEGY_HEADER_FILL = PatternFill(start_color="4A235A", end_color="4A235A", fill_type="solid")
_STRATEGY_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")

# ── Data cell styling ─────────────────────────────────────────────────
_DATA_FONT = Font(name="Calibri", size=10)
_DATA_ALIGNMENT = Alignment(horizontal="center", vertical="center")

# Signal colour map for technical-analysis columns
_SIGNAL_FILLS = {
    "Strong Buy":  PatternFill(start_color="006100", end_color="006100", fill_type="solid"),
    "Buy":         PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Neutral":     PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Sell":        PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "Strong Sell": PatternFill(start_color="9C0006", end_color="9C0006", fill_type="solid"),
    "N/A":         PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}
_SIGNAL_FONTS = {
    "Strong Buy":  Font(name="Calibri", size=10, color="FFFFFF", bold=True),
    "Buy":         Font(name="Calibri", size=10, color="006100"),
    "Neutral":     Font(name="Calibri", size=10, color="9C6500"),
    "Sell":        Font(name="Calibri", size=10, color="9C0006"),
    "Strong Sell": Font(name="Calibri", size=10, color="FFFFFF", bold=True),
    "N/A":         Font(name="Calibri", size=10, color="808080", italic=True),
}

# Strategy direction styling
_DIRECTION_FILLS = {
    "LONG":  PatternFill(start_color="006100", end_color="006100", fill_type="solid"),
    "SHORT": PatternFill(start_color="9C0006", end_color="9C0006", fill_type="solid"),
    "FLAT":  PatternFill(start_color="808080", end_color="808080", fill_type="solid"),
}
_DIRECTION_FONTS = {
    "LONG":  Font(name="Calibri", size=10, color="FFFFFF", bold=True),
    "SHORT": Font(name="Calibri", size=10, color="FFFFFF", bold=True),
    "FLAT":  Font(name="Calibri", size=10, color="FFFFFF", italic=True),
}

# Confidence styling
_CONFIDENCE_FILLS = {
    "HIGH":   PatternFill(start_color="006100", end_color="006100", fill_type="solid"),
    "MEDIUM": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "LOW":    PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}
_CONFIDENCE_FONTS = {
    "HIGH":   Font(name="Calibri", size=10, color="FFFFFF", bold=True),
    "MEDIUM": Font(name="Calibri", size=10, color="9C6500"),
    "LOW":    Font(name="Calibri", size=10, color="808080", italic=True),
}

# Optimal column widths
_COLUMN_WIDTHS = {
    "Date-Time":       22,
    "Price":           14,
    "1 min":           14,
    "5 min":           14,
    "15 min":          14,
    "30 min":          14,
    "Hourly":          14,
    "5 Hours":         14,
    "Daily":           14,
    "Weekly":          14,
    "Monthly":         14,
    # Strategy columns
    "Signal":          12,
    "Bias":            38,
    "Confidence":      14,
    "Entry":           14,
    "Stop Loss":       14,
    "Take Profit":     14,
    "Risk ($)":        12,
    "Position (lots)": 16,
    "Reason":          60,
}


class ExcelStorage:
    """Manages daily Excel files for storing scraped XAU/USD data."""

    def __init__(self) -> None:
        self._current_date: Optional[date] = None
        self._current_file: Optional[Path] = None
        self._wb: Optional[Workbook] = None
        self._ws = None
        self._row_count: int = 0
        self._save_counter: int = 0
        self._SAVE_EVERY_N_ROWS: int = 10  # Flush to disk periodically

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def append_row(self, row: Dict[str, Any]) -> None:
        """
        Append a validated data row to the current day's Excel file.

        Handles file creation, day rollover, formatting, and periodic saves.
        """
        today = date.today()

        # Day rollover: close current file, open new one
        if self._current_date != today:
            self._rotate_file(today)

        # Write data cells
        row_idx = self._ws.max_row + 1
        for col_idx, col_name in enumerate(ALL_EXCEL_COLUMNS, start=1):
            cell = self._ws.cell(row=row_idx, column=col_idx)
            value = row.get(col_name)

            if col_name == "Date-Time" and isinstance(value, datetime):
                cell.value = value.replace(tzinfo=None)
                cell.number_format = "YYYY-MM-DD HH:mm:ss"
                cell.font = _DATA_FONT
                cell.alignment = Alignment(horizontal="left", vertical="center")

            elif col_name == "Price":
                cell.value = value
                cell.number_format = "#,##0.00"
                cell.font = Font(name="Calibri", size=10, bold=True)
                cell.alignment = _DATA_ALIGNMENT

            elif col_name in STRATEGY_COLUMNS:
                # ── Strategy columns ──────────────────────────────
                self._format_strategy_cell(cell, col_name, value)

            else:
                # ── Technical signal columns ──────────────────────
                cell.value = value if value else "N/A"
                signal_key = str(cell.value)
                cell.font = _SIGNAL_FONTS.get(signal_key, _DATA_FONT)
                cell.fill = _SIGNAL_FILLS.get(signal_key, PatternFill())
                cell.alignment = _DATA_ALIGNMENT

        self._row_count += 1
        self._save_counter += 1

        # Periodic save to prevent data loss
        if self._save_counter >= self._SAVE_EVERY_N_ROWS:
            self._safe_save()
            self._save_counter = 0

        log.debug(f"Row #{self._row_count} appended to {self._current_file.name}")

    def flush(self) -> None:
        """Force save the current workbook to disk."""
        if self._wb and self._current_file:
            self._safe_save()
            log.info(f"Flushed {self._current_file.name} ({self._row_count} rows)")

    def close(self) -> None:
        """Save and close the workbook."""
        if self._wb:
            self._safe_save()
            try:
                self._wb.close()
            except Exception:
                pass
            log.info(f"Closed {self._current_file.name if self._current_file else 'workbook'}")

    @property
    def current_file(self) -> Optional[Path]:
        return self._current_file

    @property
    def row_count(self) -> int:
        return self._row_count

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _rotate_file(self, new_date: date) -> None:
        """Close current file (if any) and open/create the file for new_date."""
        # Save & close existing
        if self._wb:
            self._safe_save()
            try:
                self._wb.close()
            except Exception:
                pass

        self._current_date = new_date
        date_str = new_date.strftime("%Y-%m-%d")
        self._current_file = get_excel_filename(date_str)

        if self._current_file.exists():
            # Resume existing file
            log.info(f"Resuming existing file: {self._current_file.name}")
            self._wb = load_workbook(str(self._current_file))
            self._ws = self._wb.active
            self._row_count = max(0, self._ws.max_row - 1)  # Subtract header row
        else:
            # Create new file with formatted header
            log.info(f"Creating new file: {self._current_file.name}")
            self._wb = Workbook()
            self._ws = self._wb.active
            self._ws.title = "XAU_USD_Data"
            self._write_header()
            self._row_count = 0
            self._safe_save()

        self._save_counter = 0

    def _write_header(self) -> None:
        """Write formatted column headers to the first row."""
        for col_idx, col_name in enumerate(ALL_EXCEL_COLUMNS, start=1):
            cell = self._ws.cell(row=1, column=col_idx, value=col_name)
            cell.alignment = _HEADER_ALIGNMENT
            cell.border = _HEADER_BORDER

            # Use distinct colour for strategy section
            if col_name in STRATEGY_COLUMNS:
                cell.font = _STRATEGY_HEADER_FONT
                cell.fill = _STRATEGY_HEADER_FILL
            else:
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL

            # Set column width
            col_letter = get_column_letter(col_idx)
            self._ws.column_dimensions[col_letter].width = _COLUMN_WIDTHS.get(col_name, 14)

        # Freeze header row
        self._ws.freeze_panes = "A2"

        # Enable auto-filter over all columns
        last_col = get_column_letter(len(ALL_EXCEL_COLUMNS))
        self._ws.auto_filter.ref = f"A1:{last_col}1"

    def _format_strategy_cell(
        self, cell, col_name: str, value: Any
    ) -> None:
        """Apply formatting to a strategy-output cell."""
        cell.alignment = _DATA_ALIGNMENT

        if col_name == "Signal":
            cell.value = value if value else "FLAT"
            key = str(cell.value)
            cell.font = _DIRECTION_FONTS.get(key, _DATA_FONT)
            cell.fill = _DIRECTION_FILLS.get(key, PatternFill())

        elif col_name == "Confidence":
            cell.value = value if value else "LOW"
            key = str(cell.value)
            cell.font = _CONFIDENCE_FONTS.get(key, _DATA_FONT)
            cell.fill = _CONFIDENCE_FILLS.get(key, PatternFill())

        elif col_name in ("Entry", "Stop Loss", "Take Profit"):
            cell.value = value
            cell.number_format = "#,##0.00"
            cell.font = _DATA_FONT

        elif col_name == "Risk ($)":
            cell.value = value
            cell.number_format = "#,##0.00"
            cell.font = _DATA_FONT

        elif col_name == "Position (lots)":
            cell.value = value
            cell.number_format = "0.0000"
            cell.font = _DATA_FONT

        elif col_name == "Bias":
            cell.value = value if value else ""
            cell.font = _DATA_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        elif col_name == "Reason":
            cell.value = value if value else ""
            cell.font = Font(name="Calibri", size=9, color="444444")
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        else:
            cell.value = value
            cell.font = _DATA_FONT

    def _safe_save(self) -> None:
        """
        Save workbook using a temp-file strategy to prevent corruption.

        Writes to a temporary file first, then atomically replaces the target.
        A .bak copy is kept so that a crash between delete and rename cannot
        destroy the only copy of the data.
        """
        if not self._wb or not self._current_file:
            return

        tmp_path: Optional[Path] = None
        try:
            # Write to temp file in the same directory
            tmp_fd = tempfile.NamedTemporaryFile(
                dir=str(DATA_DIR),
                suffix=".xlsx",
                delete=False,
            )
            tmp_path = Path(tmp_fd.name)
            tmp_fd.close()

            self._wb.save(str(tmp_path))

            # Keep a .bak copy so we never have zero copies on disk
            bak_path = self._current_file.with_suffix(".xlsx.bak")
            if self._current_file.exists():
                try:
                    shutil.copy2(str(self._current_file), str(bak_path))
                except Exception:
                    pass  # Best-effort backup

            # os.replace() is the closest to atomic on Windows — it
            # overwrites the destination in a single filesystem call.
            os.replace(str(tmp_path), str(self._current_file))
            tmp_path = None  # Successfully moved, nothing to clean up

        except PermissionError:
            log.warning(
                f"Cannot save — file may be open in Excel. "
                f"Data is buffered and will be saved on next attempt."
            )
            if tmp_path is not None:
                try:
                    tmp_path.unlink()
                except Exception:
                    pass
        except Exception as e:
            log.error(f"Error saving workbook: {e}")
            if tmp_path is not None:
                try:
                    tmp_path.unlink()
                except Exception:
                    pass

