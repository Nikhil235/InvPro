"""
journal.py -- Trade journal with SQLite persistence and exports.

Provides export utilities for CSV, Excel, and JSON.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Dict, List, Optional
from datetime import date

from paper_trading.config.settings import OUTPUT_DIR
from paper_trading.core.order_manager import ClosedTrade, RejectedOrder
from paper_trading.utils.logger import get_logger
from core.database import Database

log = get_logger("journal")

_CSV_COLUMNS = [
    "trade_id", "side", "open_time", "close_time", "entry_price",
    "fill_price", "exit_price", "stop_loss", "take_profit", "exit_reason",
    "lots", "gross_pnl", "commission", "net_pnl", "risk_amount",
    "rr_achieved", "confidence", "bias", "balance_after", "reason", "session_id"
]

class TradeJournal:
    def __init__(self, db: Database, session_id: str) -> None:
        self._db = db
        self._session_id = session_id
        log.info(f"Journal initialised for session {session_id}")

    def record_trade(self, trade: ClosedTrade) -> None:
        """Record a closed trade. The broker already persists trades, so this logs and updates external files if needed."""
        log.info(f"Trade TRD-{trade.trade_id} logged to journal: {trade.side.value} lots={trade.lots} price={trade.exit_price} net_pnl=${trade.net_pnl}")

    def record_event(self, event: Dict[str, Any]) -> None:
        """Record an arbitrary event to the events_log table."""
        from datetime import datetime, timezone
        timestamp = event.get("timestamp") or datetime.now(timezone.utc).isoformat()
        event_type = event.get("type") or event.get("event_type") or "UNKNOWN"
        
        # Ensure the event itself has type/timestamp if we modified/added them
        event_copy = dict(event)
        if "type" not in event_copy:
            event_copy["type"] = event_type
        if "timestamp" not in event_copy:
            event_copy["timestamp"] = timestamp
            
        sql = "INSERT INTO events_log (timestamp, event_type, payload, session_id) VALUES (?, ?, ?, ?)"
        self._db.execute(sql, (
            timestamp,
            event_type,
            json.dumps(event_copy),
            self._session_id
        ))

    def export_excel(self, file_path: Optional[Path] = None) -> None:
        trades = self._get_trades()
        if not trades:
            log.info("No trades to export to Excel.")
            return

        if not file_path:
            file_path = OUTPUT_DIR / f"journal_{self._session_id}.xlsx"

        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            rows = [t.to_dict() for t in trades]
            df = pd.DataFrame(rows)

            available = [c for c in _CSV_COLUMNS if c in df.columns]
            extra = [c for c in df.columns if c not in _CSV_COLUMNS]
            df = df[available + extra]

            wb = Workbook()
            ws = wb.active
            ws.title = "Trade_Journal"

            header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            for col_idx, col_name in enumerate(df.columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col_name)) + 4, 14)

            win_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            loss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
                for col_idx, val in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.alignment = Alignment(horizontal="center")

                net_pnl = row.get("net_pnl", 0)
                fill = win_fill if net_pnl > 0 else loss_fill
                for col_idx in range(1, len(df.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}1"

            wb.save(str(file_path))
            log.info(f"Excel export saved: {file_path}")

        except ImportError as e:
            log.warning(f"Could not export Excel (missing dependency): {e}")
        except Exception as e:
            log.error(f"Error exporting Excel: {e}")

    def export_csv(self, file_path: Optional[Path] = None) -> None:
        trades = self._get_trades()
        if not trades:
            return
        
        if not file_path:
            file_path = OUTPUT_DIR / f"journal_{self._session_id}.csv"

        try:
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=_CSV_COLUMNS, extrasaction="ignore")
                writer.writeheader()
                for t in trades:
                    writer.writerow(t.to_dict())
            log.info(f"CSV export saved: {file_path}")
        except Exception as e:
            log.error(f"Error exporting CSV: {e}")

    def export_json(self, file_path: Optional[Path] = None) -> None:
        trades = self._get_trades()
        if not trades:
            return
            
        if not file_path:
            file_path = OUTPUT_DIR / f"journal_{self._session_id}.json"

        try:
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump([t.to_dict() for t in trades], f, indent=2, default=str)
            log.info(f"JSON export saved: {file_path}")
        except Exception as e:
            log.error(f"Error exporting JSON: {e}")

    def _get_trades(self) -> List[ClosedTrade]:
        rows = self._db.fetchall("SELECT * FROM trades WHERE session_id = ? ORDER BY close_time ASC", (self._session_id,))
        return [ClosedTrade.from_db_row(row) for row in rows]

    @property
    def trades(self) -> List[ClosedTrade]:
        """Get all closed trades for this session."""
        return self._get_trades()

    @property
    def csv_path(self) -> Path:
        """Get the path to the CSV journal file."""
        return OUTPUT_DIR / f"journal_{self._session_id}.csv"

    @property
    def xlsx_path(self) -> Path:
        """Get the path to the Excel journal file."""
        return OUTPUT_DIR / f"journal_{self._session_id}.xlsx"

    @property
    def jsonl_path(self) -> Path:
        """Get the path to the JSON/Event Log journal file."""
        return OUTPUT_DIR / f"journal_{self._session_id}.json"
