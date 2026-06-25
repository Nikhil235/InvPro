"""Tests for core.storage — Excel file creation, writing, rotation, safe save."""

from __future__ import annotations

import pytest
import tempfile
import shutil
from datetime import datetime, date, timezone
from pathlib import Path
from unittest.mock import patch

from core.storage import ExcelStorage


@pytest.fixture
def tmp_data_dir(tmp_path):
    """Provide a temporary data directory and patch settings to use it."""
    with patch("core.storage.DATA_DIR", tmp_path), \
         patch("core.storage.get_excel_filename", lambda d: tmp_path / f"xau_usd_{d}.xlsx"):
        yield tmp_path


@pytest.fixture
def storage(tmp_data_dir):
    s = ExcelStorage()
    yield s
    s.close()


def _make_row(**overrides):
    defaults = {
        "Date-Time": datetime(2026, 6, 23, 12, 0, 0, tzinfo=timezone.utc),
        "Price": 4100.00,
        "1 min": "Buy",
        "5 min": "Neutral",
        "15 min": "Sell",
        "30 min": "Buy",
        "Hourly": "Neutral",
        "5 Hours": "Sell",
        "Daily": "Strong Buy",
        "Weekly": "Buy",
        "Monthly": "Strong Buy",
    }
    defaults.update(overrides)
    return defaults


class TestExcelStorage:
    def test_creates_file_on_first_append(self, storage, tmp_data_dir):
        row = _make_row()
        storage.append_row(row)
        assert storage.current_file is not None
        assert storage.current_file.exists()

    def test_row_count_increments(self, storage):
        storage.append_row(_make_row())
        storage.append_row(_make_row(Price=4101.0))
        assert storage.row_count == 2

    def test_header_written_correctly(self, storage, tmp_data_dir):
        from openpyxl import load_workbook
        storage.append_row(_make_row())
        storage.flush()

        wb = load_workbook(str(storage.current_file))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "Date-Time" in headers
        assert "Price" in headers
        assert "Signal" in headers
        assert "Stop Loss" in headers
        wb.close()

    def test_flush_writes_to_disk(self, storage):
        storage.append_row(_make_row())
        storage.flush()
        assert storage.current_file.exists()
        assert storage.current_file.stat().st_size > 0

    def test_safe_save_creates_bak_file(self, storage, tmp_data_dir):
        storage.append_row(_make_row())
        storage.flush()  # first save
        storage.append_row(_make_row(Price=4102.0))
        storage.flush()  # second save should create .bak
        bak_file = storage.current_file.with_suffix(".xlsx.bak")
        assert bak_file.exists()

    def test_close_saves_and_releases(self, storage):
        storage.append_row(_make_row())
        file_path = storage.current_file
        storage.close()
        assert file_path.exists()
