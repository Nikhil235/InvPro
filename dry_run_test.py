"""
dry_run_test.py — Quick verification test for the pipeline.

Runs 3 scrape cycles to confirm browser launch, data extraction,
parsing, validation, and Excel storage all work correctly.
"""

import sys
import os
import time

# Fix Windows console encoding
if sys.platform == "win32":
    os.environ.setdefault("PYTHONIOENCODING", "utf-8")
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass



from config.settings import EXCEL_COLUMNS
from core.scraper import GoldScraper
from core.parser import parse_snapshot
from core.validator import DataValidator
from core.storage import ExcelStorage
from utils.logger import get_logger

log = get_logger("dry_run")

def main():
    print("=" * 60)
    print("  DRY RUN TEST -- 3 scrape cycles")
    print("=" * 60)

    scraper = GoldScraper()
    validator = DataValidator()
    storage = ExcelStorage()

    try:
        # Launch browser
        print("\n[1/4] Launching browser...")
        scraper.start()
        print("  [OK] Browser launched successfully")

        # Run 3 cycles
        for cycle in range(1, 4):
            print(f"\n[Cycle {cycle}/3] Scraping...")

            # Scrape
            raw = scraper.read_snapshot()
            print(f"  Raw data: price={raw.get('price')}")
            for key, val in raw.items():
                if key != "price":
                    print(f"    {key}: {val}")

            # Parse
            row = parse_snapshot(raw)
            print(f"  Parsed: price={row.get('Price')}, time={row.get('Date-Time')}")

            # Validate
            is_valid, reason = validator.validate(row)
            print(f"  Valid: {is_valid} ({reason})")

            # Store
            if is_valid:
                storage.append_row(row)
                print(f"  [OK] Stored to {storage.current_file.name} (row #{storage.row_count})")
            else:
                print(f"  [SKIP] Skipped: {reason}")

            if cycle < 3:
                print("  Waiting 5s...")
                time.sleep(5)

        # Flush and verify
        storage.flush()
        print(f"\n[4/4] Final file: {storage.current_file}")
        print(f"  Total rows: {storage.row_count}")

        # Read back and verify
        from openpyxl import load_workbook
        wb = load_workbook(str(storage.current_file))
        ws = wb.active
        print(f"  Sheet name: {ws.title}")
        print(f"  Headers: {[cell.value for cell in ws[1]]}")
        if ws.max_row >= 2:
            print(f"  First data row: {[cell.value for cell in ws[2]]}")
        wb.close()

        print("\n" + "=" * 60)
        print("  [PASS] DRY RUN PASSED -- Pipeline is working!")
        print("=" * 60)

    except Exception as e:
        print(f"\n  [FAIL] DRY RUN FAILED: {e}")
        import traceback
        traceback.print_exc()
    finally:
        storage.close()
        scraper.stop()


if __name__ == "__main__":
    main()
